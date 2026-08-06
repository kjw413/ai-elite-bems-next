"""절감테마_양식.xlsx 파싱, 미리보기, 트랜잭션 일괄 등록.

양식의 연누계(E/F)는 월별 값에서 계산되는 표시 열이므로 저장에 사용하지 않는다.
신규 테마는 화면 기본 필터에 바로 보이도록 ``ongoing`` 상태로 등록한다. 같은
(공장, 연도, 테마명)이 이미 있으면 수기 관리 항목(상태·시행월·담당·투자비·메모)은
보존하고, 양식이 책임지는 에너지원·분류·월별 계획/실적만 갱신한다.
"""
from __future__ import annotations

import math
from collections import Counter
from typing import Any, BinaryIO
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException

from app.database.db_connection import managed_cursor
from app.services.audit_service import get_current_user


FACTORY_SHEETS: tuple[str, ...] = ("남양주", "김해", "광주", "논산", "경산")
ENERGY_LABEL_TO_TYPE: dict[str, str] = {"전력": "power", "연료": "fuel", "용수": "water"}
ENERGY_TYPE_TO_LABEL: dict[str, str] = {value: key for key, value in ENERGY_LABEL_TO_TYPE.items()}
CATEGORIES: tuple[str, ...] = ("설비교체", "운전개선", "공정개선", "누설저감", "계약변경", "기타")
NEW_THEME_STATUS = "ongoing"

EXPECTED_HEADERS: tuple[str, ...] = (
    "No.", "테마명", "에너지원", "절감유형", "연누계(계획)", "연누계(실적)",
    "1월(계획)", "1월(실적)", "2월(계획)", "2월(실적)",
    "3월(계획)", "3월(실적)", "4월(계획)", "4월(실적)",
    "5월(계획)", "5월(실적)", "6월(계획)", "6월(실적)",
    "7월(계획)", "7월(실적)", "8월(계획)", "8월(실적)",
    "9월(계획)", "9월(실적)", "10월(계획)", "10월(실적)",
    "11월(계획)", "11월(실적)", "12월(계획)", "12월(실적)",
)


class TemplateValidationError(ValueError):
    """사용자가 바로 고칠 수 있는 양식/셀 오류."""


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _number(value: Any, location: str, *, blank: float | None) -> float | None:
    if value is None or (isinstance(value, str) and not value.strip()):
        return blank
    if isinstance(value, bool):
        raise TemplateValidationError(f"{location}: 계획·실적은 숫자로 입력하세요.")
    candidate: Any = value
    if isinstance(value, str):
        candidate = value.strip().replace(",", "")
        if candidate.startswith("="):
            raise TemplateValidationError(f"{location}: 월별 계획·실적에는 수식 대신 숫자를 입력하세요.")
    try:
        parsed = float(candidate)
    except (TypeError, ValueError) as exc:
        raise TemplateValidationError(f"{location}: 계획·실적은 숫자로 입력하세요.") from exc
    if not math.isfinite(parsed):
        raise TemplateValidationError(f"{location}: 유한한 숫자를 입력하세요.")
    return parsed


def derive_start_ym(records: list[dict[str, Any]], year: int) -> str | None:
    """계획량 또는 실적량이 처음 기입된 월 → 'YYYY-MM' (시행월).

    양식에 시행월 열이 없어 월별 값에서 도출한다(2026-08 결정). 이 값은 원단위
    전후 비교 검증의 분기점이므로, 없으면 그 테마는 영영 '판정 보류'가 된다.

    주의 — 연초부터 계획이 깔린 테마는 시행월이 1월로 잡혀 '시행 전' 구간이
    통째로 전년도가 된다. 연중 시행 건은 시행 전 월의 계획을 0(또는 공란)으로
    두어야 전후 비교가 의도대로 나뉜다. 관리자 화면에서 개별 수정할 수 있다.
    """
    for record in records:
        planned = float(record.get("planned_qty") or 0.0)
        actual = record.get("actual_qty")
        if planned != 0.0 or actual is not None:
            return f"{int(year):04d}-{int(record['month']):02d}"
    return None


def _validate_headers(sheet: Any) -> None:
    actual = tuple(_text(sheet.cell(1, column).value) for column in range(1, len(EXPECTED_HEADERS) + 1))
    for column, (found, expected) in enumerate(zip(actual, EXPECTED_HEADERS), start=1):
        if found != expected:
            coordinate = sheet.cell(1, column).coordinate
            raise TemplateValidationError(
                f"{sheet.title}!{coordinate}: 열 제목이 '{expected}'이어야 합니다. "
                "제공된 절감테마 양식을 사용하세요."
            )


def parse_template(source: BinaryIO, year: int) -> list[dict[str, Any]]:
    """절감 테마 엑셀을 DB 저장용 행 목록으로 변환한다."""
    if not 2000 <= int(year) <= 2100:
        raise TemplateValidationError("등록 연도는 2000~2100 사이여야 합니다.")
    try:
        workbook = load_workbook(source, read_only=True, data_only=False)
    except (BadZipFile, InvalidFileException, OSError, ValueError, KeyError) as exc:
        raise TemplateValidationError("올바른 .xlsx 파일이 아니거나 파일이 손상되었습니다.") from exc

    found_sheets = [name for name in FACTORY_SHEETS if name in workbook.sheetnames]
    if not found_sheets:
        raise TemplateValidationError(
            "남양주·김해·광주·논산·경산 시트가 없습니다. 제공된 절감테마 양식을 사용하세요."
        )

    themes: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    try:
        for sheet_name in found_sheets:
            sheet = workbook[sheet_name]
            _validate_headers(sheet)
            for row_number, cells in enumerate(
                sheet.iter_rows(min_row=2, max_col=len(EXPECTED_HEADERS)), start=2,
            ):
                # 번호와 연누계 수식은 행 유무 판정에서도 제외한다. 원본 양식은 빈 2행의
                # E/F에 합계 수식만 들어 있으므로 이 두 셀 때문에 빈 테마가 생기면 안 된다.
                input_cells = (*cells[1:4], *cells[6:30])
                if not any(_text(cell.value) for cell in input_cells):
                    continue

                title = _text(cells[1].value)
                if not title:
                    raise TemplateValidationError(f"{sheet_name}!B{row_number}: 테마명을 입력하세요.")
                if len(title) > 255:
                    raise TemplateValidationError(f"{sheet_name}!B{row_number}: 테마명은 255자 이하여야 합니다.")

                energy_label = _text(cells[2].value)
                energy_type = ENERGY_LABEL_TO_TYPE.get(energy_label)
                if energy_type is None:
                    raise TemplateValidationError(
                        f"{sheet_name}!C{row_number}: 에너지원은 전력/연료/용수 중 하나여야 합니다."
                    )

                category = _text(cells[3].value)
                if category not in CATEGORIES:
                    raise TemplateValidationError(
                        f"{sheet_name}!D{row_number}: 절감유형은 {'/'.join(CATEGORIES)} 중 하나여야 합니다."
                    )

                duplicate_key = (sheet_name, title.casefold())
                if duplicate_key in seen:
                    raise TemplateValidationError(
                        f"{sheet_name}!B{row_number}: 같은 공장 시트에 동일한 테마명이 중복되었습니다."
                    )
                seen.add(duplicate_key)

                records: list[dict[str, Any]] = []
                for month in range(1, 13):
                    planned_index = 6 + (month - 1) * 2
                    actual_index = planned_index + 1
                    planned_cell = cells[planned_index]
                    actual_cell = cells[actual_index]
                    records.append({
                        "month": month,
                        "planned_qty": _number(
                            planned_cell.value,
                            f"{sheet_name}!{get_column_letter(planned_index + 1)}{row_number}",
                            blank=0.0,
                        ),
                        "actual_qty": _number(
                            actual_cell.value,
                            f"{sheet_name}!{get_column_letter(actual_index + 1)}{row_number}",
                            blank=None,
                        ),
                    })

                themes.append({
                    "factory": sheet_name,
                    "year": int(year),
                    "title": title,
                    "energy_type": energy_type,
                    "category": category,
                    "status": NEW_THEME_STATUS,
                    "start_ym": derive_start_ym(records, int(year)),
                    "records": records,
                    "source_row": row_number,
                })
    finally:
        workbook.close()

    if not themes:
        raise TemplateValidationError("등록할 절감 테마가 없습니다.")
    return themes


def _existing_theme_ids(themes: list[dict[str, Any]]) -> dict[tuple[str, str], int]:
    year = int(themes[0]["year"])
    factories = sorted({str(theme["factory"]) for theme in themes})
    placeholders = ",".join(["%s"] * len(factories))
    with managed_cursor(dictionary=True) as (_conn, cursor):
        cursor.execute(
            f"SELECT id, factory, title FROM savings_theme WHERE year=%s AND factory IN ({placeholders})",
            (year, *factories),
        )
        rows = cursor.fetchall()
    return {(str(row["factory"]), str(row["title"]).casefold()): int(row["id"]) for row in rows}


def preview_import(themes: list[dict[str, Any]]) -> dict[str, Any]:
    existing = _existing_theme_ids(themes)
    factory_counts = Counter(str(theme["factory"]) for theme in themes)
    items = []
    new_count = 0
    update_count = 0
    record_values = 0
    for theme in themes:
        key = (str(theme["factory"]), str(theme["title"]).casefold())
        action = "update" if key in existing else "new"
        if action == "new":
            new_count += 1
        else:
            update_count += 1
        record_values += sum(
            1 for record in theme["records"]
            if float(record["planned_qty"] or 0.0) != 0.0 or record["actual_qty"] is not None
        )
        actual_values = [
            float(record["actual_qty"])
            for record in theme["records"]
            if record["actual_qty"] is not None
        ]
        items.append({
            "factory": theme["factory"],
            "sourceRow": theme["source_row"],
            "title": theme["title"],
            "energyType": theme["energy_type"],
            "energyLabel": ENERGY_TYPE_TO_LABEL[str(theme["energy_type"])],
            "category": theme["category"],
            "action": action,
            # 도출된 시행월 — 검증 분기점이라 반영 전에 확인할 수 있게 내려보낸다.
            # 기존 테마에 이미 시행월이 있으면 그 값이 유지되므로(COALESCE) 참고값이다.
            "startYm": theme.get("start_ym"),
            "plannedTotal": sum(float(record["planned_qty"] or 0.0) for record in theme["records"]),
            "actualTotal": sum(actual_values) if actual_values else None,
            "months": [
                {
                    "month": int(record["month"]),
                    "plannedQty": float(record["planned_qty"] or 0.0),
                    "actualQty": (
                        float(record["actual_qty"])
                        if record["actual_qty"] is not None else None
                    ),
                }
                for record in theme["records"]
            ],
        })
    return {
        "success": True,
        "year": int(themes[0]["year"]),
        "totalThemes": len(themes),
        "newThemes": new_count,
        "existingThemes": update_count,
        "recordValues": record_values,
        "byFactory": [{"factory": factory, "themes": factory_counts[factory]} for factory in FACTORY_SHEETS if factory_counts[factory]],
        "items": items,
    }


def apply_import(themes: list[dict[str, Any]]) -> dict[str, int]:
    """테마와 12개월 값을 한 트랜잭션으로 반영한다."""
    user = get_current_user()
    inserted = 0
    updated = 0
    saved_records = 0
    cleared_records = 0
    with managed_cursor(admin=True) as (connection, cursor):
        try:
            for theme in themes:
                cursor.execute(
                    "SELECT id FROM savings_theme WHERE factory=%s AND year=%s AND title=%s FOR UPDATE",
                    (theme["factory"], theme["year"], theme["title"]),
                )
                rows = cursor.fetchall()
                if rows:
                    theme_id = int(rows[0]["id"] if isinstance(rows[0], dict) else rows[0][0])
                    # 시행월은 COALESCE 로 "비어 있을 때만" 채운다 — 관리자가 화면에서
                    # 직접 지정한 값은 양식이 덮지 않고, 아직 NULL 인 건만 도출값으로
                    # 메운다(수기 관리 항목 보존 원칙 + 기존 NULL 건 백필을 동시에).
                    cursor.execute(
                        """
                        UPDATE savings_theme
                        SET energy_type=%s, category=%s,
                            start_ym=COALESCE(start_ym, %s), changed_by=%s
                        WHERE id=%s
                        """,
                        (
                            theme["energy_type"], theme["category"],
                            theme.get("start_ym"), user, theme_id,
                        ),
                    )
                    updated += 1
                else:
                    cursor.execute(
                        """
                        INSERT INTO savings_theme
                            (factory, year, title, energy_type, category, status,
                             start_ym, owner, invest_amount, note, changed_by)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, NULL, NULL, NULL, %s)
                        """,
                        (
                            theme["factory"], theme["year"], theme["title"],
                            theme["energy_type"], theme["category"], theme["status"],
                            theme.get("start_ym"), user,
                        ),
                    )
                    theme_id = int(cursor.lastrowid)
                    inserted += 1

                for record in theme["records"]:
                    planned = float(record["planned_qty"] or 0.0)
                    actual = record["actual_qty"]
                    if planned == 0.0 and actual is None:
                        cursor.execute(
                            "DELETE FROM savings_record WHERE theme_id=%s AND year=%s AND month=%s",
                            (theme_id, theme["year"], record["month"]),
                        )
                        cleared_records += int(cursor.rowcount or 0)
                    else:
                        cursor.execute(
                            """
                            INSERT INTO savings_record
                                (theme_id, year, month, planned_qty, actual_qty, changed_by)
                            VALUES (%s, %s, %s, %s, %s, %s)
                            ON DUPLICATE KEY UPDATE
                                planned_qty=VALUES(planned_qty),
                                actual_qty=VALUES(actual_qty),
                                changed_by=VALUES(changed_by)
                            """,
                            (theme_id, theme["year"], record["month"], planned, actual, user),
                        )
                        saved_records += 1
            connection.commit()
        except Exception:
            connection.rollback()
            raise
    return {
        "insertedThemes": inserted,
        "updatedThemes": updated,
        "savedRecords": saved_records,
        "clearedRecords": cleared_records,
    }
