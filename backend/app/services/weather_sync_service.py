# 이 파일은 기상청 API Hub(ASOS)에서 날씨 데이터를 자동 수집하여 DB_weather.xlsx를 최신으로 유지합니다.
from __future__ import annotations

import os
import logging
from datetime import date, timedelta
from pathlib import Path
from typing import Any

import pandas as pd
import requests

# 공장 시트명 → ASOS 지점 코드 매핑 (v5_common 과 단일 정의 공유)
from app.services.v5_common import PATH_WEATHER, STATION_CODE_MAP
from app.utils.file_io import atomic_save_workbook, exclusive_file_lock
from app.utils.tls import requests_verify

# v5 모델 학습 시작일과 동기화하여 신규 수집 범위를 결정합니다.
# (modeling_v5.1.TRAIN_START / v5_retrain_worker.TRAIN_START 와 일치)
WEATHER_TRAIN_START = date(2021, 1, 1)

logger = logging.getLogger(__name__)

# ── 경로 및 상수 ────────────────────────────────────────────────
_ROOT = Path(__file__).resolve().parents[2]
WEATHER_EXCEL_LOCK_PATH = PATH_WEATHER.with_suffix(PATH_WEATHER.suffix + ".lock")

# 기상청 ASOS 일자료 API 엔드포인트
KMA_ASOS_URL = "https://apihub.kma.go.kr/api/typ01/url/kma_sfcdd3.php"

# 기상청 ASOS 일자료 실제 응답 포맷 (help=1 기준, 0-based index)
# TM(0) STN(1) WS_AVG(2) WR_DAY(3) WD_MAX(4) WS_MAX(5) WS_MAX_TM(6) WD_INS(7)
# WS_INS(8) WS_INS_TM(9) TA_AVG(10) TA_MAX(11) TA_MAX_TM(12) TA_MIN(13) TA_MIN_TM(14)
# TD_AVG(15) TS_AVG(16) TG_MIN(17) HM_AVG(18) HM_MIN(19) HM_MIN_TM(20)
# PV_AVG(21) EV_S(22) EV_L(23) FG_DUR(24) PA_AVG(25) PS_AVG(26)
# PS_MAX(27) PS_MAX_TM(28) PS_MIN(29) PS_MIN_TM(30) CA_TOT(31)
# SS_DAY(32) SS_DUR(33) SS_CMB(34) SI_DAY(35) ... RN_DAY(38)
_FIELD_INDICES = {
    "TM":     0,   # 날짜 YYYYMMDD
    "TA_AVG": 10,  # 평균기온
    "WR_DAY": 3,   # 일강수량(m 단위 주의)
    "RN_DAY": 38,  # 일강수량(mm) - 우선 사용
    "HM_AVG": 18,  # 평균 상대습도 %
    "SI_DAY": 35,  # 합계 일사량 MJ/m2
    "SS_DAY": 32,  # 합계 일조시간 hr
}


def _get_kma_api_key() -> str | None:
    """환경변수에서 기상청 API 키를 읽습니다."""
    key = os.environ.get("KMA_API_KEY", "").strip()
    if not key:
        # dotenv 미적용 환경을 위한 폴백: .env 파일 직접 읽기
        env_path = _ROOT / ".env"
        if env_path.exists():
            for line in env_path.read_text(encoding="utf-8").splitlines():
                if line.startswith("KMA_API_KEY="):
                    key = line.split("=", 1)[1].strip()
                    break
    return key or None


def fetch_kma_asos(station_code: int, date_from: date, date_to: date) -> pd.DataFrame:
    """
    기상청 ASOS 일자료 API를 호출하여 날씨 DataFrame을 반환합니다.

    Returns
    -------
    DataFrame with columns:
      [\"일시\", \"평균기온(°C)\", \"일강수량(mm)\", \"평균 상대습도(%)\", \"합계 일사량(MJ/m2)\", \"합계 일조시간(hr)\"]
    실패 시 빈 DataFrame 반환.
    """
    api_key = _get_kma_api_key()
    if not api_key:
        logger.warning("[weather_sync] KMA_API_KEY가 설정되지 않아 날씨 데이터를 수집할 수 없습니다.")
        return pd.DataFrame()

    params = {
        "tm1": date_from.strftime("%Y%m%d"),
        "tm2": date_to.strftime("%Y%m%d"),
        "stn": station_code,
        "help": 1,
        "authKey": api_key,
    }

    try:
        resp = requests.get(KMA_ASOS_URL, params=params, timeout=15, verify=requests_verify())
        resp.raise_for_status()
    except requests.RequestException as e:
        logger.error(f"[weather_sync] API 호출 실패 (stn={station_code}): {e}")
        return pd.DataFrame()

    parsed = _parse_kma_response(resp.text, station_code)
    if parsed.empty:
        logger.warning(f"[weather_sync] stn={station_code} 파싱 결과 0행. 응답 일부: {resp.text[:200]}")
    return parsed



def _parse_kma_response(text: str, station_code: int) -> pd.DataFrame:
    """
    기상청 API 텍스트 응답을 DataFrame으로 파싱합니다.
    포맷: 공백 구분, '#' 시작 줄은 주석/헤더.
    실제 필드 순서 (0-based): TM(0) STN(1) ... TA_AVG(10) HM_AVG(18)
      SS_DAY(32) SI_DAY(35) RN_DAY(38)
    """
    rows = []
    for line in text.splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        parts = line.split()
        # SI_DAY가 35번 인덱스이므로 최소 36개 필드 필요
        if len(parts) < 36:
            continue
        try:
            tm_str = parts[0]
            dt_val = pd.to_datetime(tm_str[:8], format="%Y%m%d", errors="coerce")
            if pd.isna(dt_val):
                continue

            def _safe(idx: int) -> float:
                """필드 1개를 안전하게 float으로 변환합니다.

                결측(미측정/응답 누락/파싱 실패) 시에는 0이 아니라 NaN을 반환합니다.
                0과 "측정 안 됨"은 의미가 다르므로 다운스트림에서 별도 보간이 필요합니다.
                """
                if idx >= len(parts):
                    return float("nan")
                val = parts[idx]
                if val is None or val == "" or val == "-":
                    return float("nan")
                try:
                    f = float(val)
                    # 기상청 결측 코드: -9.0(미측정), -99.0, -999.0 등
                    return float("nan") if f < -8 else f
                except (ValueError, TypeError):
                    return float("nan")

            # RN_DAY(38: mm) — 응답에 없으면 _safe가 NaN 반환.
            # 이후 fill_weather_gaps에서 강수량은 결측을 0(무강수)으로 처리합니다.
            rows.append({
                "일시": dt_val,
                "평균기온(°C)": _safe(10),      # TA_AVG
                "일강수량(mm)": _safe(38),       # RN_DAY
                "평균 상대습도(%)": _safe(18),   # HM_AVG
                "합계 일사량(MJ/m2)": _safe(35), # SI_DAY
                "합계 일조시간(hr)": _safe(32),  # SS_DAY
            })
        except Exception:
            continue

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)
    df["일시"] = pd.to_datetime(df["일시"])
    return df.drop_duplicates(subset=["일시"]).sort_values("일시").reset_index(drop=True)



def get_last_weather_date(station_name: str) -> date | None:
    """
    DB_weather.xlsx에서 해당 관측소 시트의 마지막 날짜를 반환합니다.
    파일/시트가 없으면 None 반환.
    """
    if not PATH_WEATHER.exists():
        return None
    try:
        df = pd.read_excel(PATH_WEATHER, sheet_name=station_name)
    except Exception:
        return None

    if df.empty or "일시" not in df.columns:
        return None

    dates = pd.to_datetime(df["일시"], errors="coerce").dropna()
    if dates.empty:
        return None

    return dates.max().date()


def sync_weather_excel(station_name: str, station_code: int) -> dict[str, Any]:
    """
    특정 관측소의 날씨 데이터를 DB_weather.xlsx에 동기화합니다.
    마지막 보유 날짜 이후 ~ 어제(T-1)까지의 누락 데이터를 API로 채웁니다.

    Returns
    -------
    dict with keys: station, added_days, last_date, error
    """
    yesterday = date.today() - timedelta(days=1)
    last_date = get_last_weather_date(station_name)

    if last_date is None:
        # 데이터가 전혀 없으면 v5 학습 시작일(WEATHER_TRAIN_START)부터 수집합니다.
        # 과거에는 yesterday-2년부터만 수집하여 학습 구간(2021~)이 결측 처리되는 결함이 있었습니다.
        fetch_from = WEATHER_TRAIN_START
        logger.info(f"[weather_sync] {station_name}: 전체 신규 수집 ({fetch_from} ~ {yesterday})")
    elif last_date >= yesterday:
        logger.info(f"[weather_sync] {station_name}: 이미 최신 상태 (last={last_date})")
        return {"station": station_name, "added_days": 0, "last_date": str(last_date), "error": None}
    else:
        fetch_from = last_date + timedelta(days=1)
        logger.info(f"[weather_sync] {station_name}: 누락 수집 ({fetch_from} ~ {yesterday})")

    # API 호출
    new_df = fetch_kma_asos(station_code, fetch_from, yesterday)
    if new_df.empty:
        return {
            "station": station_name,
            "added_days": 0,
            "last_date": str(last_date) if last_date else None,
            "error": "API에서 데이터를 가져오지 못했습니다.",
        }

    # Excel 병합 및 저장
    try:
        _append_to_excel(station_name, new_df)
    except Exception as e:
        logger.error(f"[weather_sync] {station_name} Excel 저장 실패: {e}")
        return {"station": station_name, "added_days": 0, "last_date": str(last_date), "error": str(e)}

    added = len(new_df)
    new_last = new_df["일시"].max().date()
    logger.info(f"[weather_sync] {station_name}: {added}일 추가 완료 (last={new_last})")
    return {"station": station_name, "added_days": added, "last_date": str(new_last), "error": None}


def _append_to_excel(station_name: str, new_df: pd.DataFrame) -> None:
    """
    DB_weather.xlsx의 station_name 시트에 새 데이터를 append합니다.
    중복 날짜는 제거하고 날짜순으로 정렬하여 저장합니다.
    """
    import openpyxl

    PATH_WEATHER.parent.mkdir(parents=True, exist_ok=True)

    with exclusive_file_lock(WEATHER_EXCEL_LOCK_PATH):
        sheets: dict[str, pd.DataFrame] = {}
        sheet_order: list[str] = []

        if PATH_WEATHER.exists():
            try:
                with pd.ExcelFile(PATH_WEATHER) as xls:
                    sheet_order = list(xls.sheet_names)
                    sheets = {
                        sheet: pd.read_excel(xls, sheet_name=sheet)
                        for sheet in sheet_order
                    }
            except Exception as exc:
                raise RuntimeError(f"기존 날씨 엑셀을 읽을 수 없어 덮어쓰기를 중단합니다: {PATH_WEATHER}") from exc

        existing_df = sheets.get(station_name, pd.DataFrame())
        if not existing_df.empty and "일시" in existing_df.columns:
            existing_df = existing_df.copy()
            existing_df["일시"] = pd.to_datetime(existing_df["일시"], errors="coerce")
            combined = pd.concat([existing_df, new_df], ignore_index=True)
        else:
            combined = new_df.copy()

        combined["일시"] = pd.to_datetime(combined["일시"], errors="coerce")
        # concat 순서가 [existing, new]이므로 keep="last"로 새 응답값을 우선시합니다.
        # (KMA가 D+1 잠정값을 며칠 후 검정값으로 갱신하는 경우 보정값을 받기 위함)
        sheets[station_name] = (
            combined.dropna(subset=["일시"])
            .drop_duplicates(subset=["일시"], keep="last")
            .sort_values("일시")
            .reset_index(drop=True)
        )
        if station_name not in sheet_order:
            sheet_order.append(station_name)

        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        for sheet in sheet_order:
            df_sheet = sheets.get(sheet, pd.DataFrame())
            ws = wb.create_sheet(sheet)
            headers = list(df_sheet.columns)
            ws.append(headers)
            for _, row in df_sheet.iterrows():
                row_data = []
                for col in headers:
                    val = row[col]
                    if pd.isna(val):
                        row_data.append(None)
                    elif hasattr(val, "to_pydatetime"):
                        row_data.append(val.to_pydatetime())
                    else:
                        row_data.append(val)
                ws.append(row_data)

        atomic_save_workbook(wb, PATH_WEATHER)


def sync_all_stations() -> list[dict[str, Any]]:
    """
    모든 관측소(서울/김해/이천/부여)에 대해 날씨 데이터를 동기화합니다.
    예측 실행 전 또는 관리자 수동 트리거 시 호출됩니다.

    Returns
    -------
    각 관측소 결과 dict 목록
    """
    results = []
    for station_name, station_code in STATION_CODE_MAP.items():
        result = sync_weather_excel(station_name, station_code)
        results.append(result)
    return results


def get_weather_sync_status() -> dict[str, Any]:
    """
    각 관측소의 현재 날씨 데이터 최신 날짜와 어제 대비 누락일 수를 반환합니다.
    UI 상태 표시용.
    """
    yesterday = date.today() - timedelta(days=1)
    status: dict[str, Any] = {}
    for station_name in STATION_CODE_MAP:
        last = get_last_weather_date(station_name)
        if last is None:
            missing = None
        else:
            missing = max(0, (yesterday - last).days)
        status[station_name] = {
            "last_date": str(last) if last else "없음",
            "missing_days": missing,
            "is_up_to_date": (last is not None and last >= yesterday),
        }
    return status
