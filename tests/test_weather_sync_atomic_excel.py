from __future__ import annotations

import sys
import tempfile
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.services import weather_sync_service as weather


def _check(condition: bool, message: str, failures: list[str]) -> None:
    if not condition:
        failures.append(message)


def _write_seed_workbook(path: Path) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        pd.DataFrame(
            {
                "일시": [pd.Timestamp("2026-06-01")],
                "평균기온(°C)": [20.0],
                "일강수량(mm)": [0.0],
                "평균 상대습도(%)": [55.0],
                "합계 일사량(MJ/m2)": [19.0],
                "합계 일조시간(hr)": [9.0],
            }
        ).to_excel(writer, sheet_name="서울", index=False)
        pd.DataFrame(
            {
                "일시": [pd.Timestamp("2026-06-01")],
                "평균기온(°C)": [24.0],
            }
        ).to_excel(writer, sheet_name="김해", index=False)


def main() -> int:
    failures: list[str] = []
    original_path = weather.PATH_WEATHER
    original_lock = weather.WEATHER_EXCEL_LOCK_PATH

    with tempfile.TemporaryDirectory() as tmp:
        tmp_dir = Path(tmp)
        weather_path = tmp_dir / "DB_weather.xlsx"
        weather.PATH_WEATHER = weather_path
        weather.WEATHER_EXCEL_LOCK_PATH = weather_path.with_suffix(".xlsx.lock")

        try:
            _write_seed_workbook(weather_path)
            new_df = pd.DataFrame(
                {
                    "일시": [pd.Timestamp("2026-06-01"), pd.Timestamp("2026-06-02")],
                    "평균기온(°C)": [21.5, 22.0],
                    "일강수량(mm)": [1.0, 0.0],
                    "평균 상대습도(%)": [60.0, 61.0],
                    "합계 일사량(MJ/m2)": [18.0, 20.0],
                    "합계 일조시간(hr)": [8.5, 9.5],
                }
            )

            weather._append_to_excel("서울", new_df)

            wb = load_workbook(weather_path, read_only=True, data_only=True)
            try:
                _check("서울" in wb.sheetnames, "updated station sheet exists", failures)
                _check("김해" in wb.sheetnames, "other station sheet is preserved", failures)
            finally:
                wb.close()

            seoul = pd.read_excel(weather_path, sheet_name="서울")
            gimhae = pd.read_excel(weather_path, sheet_name="김해")
            _check(len(seoul) == 2, "duplicate dates are deduplicated", failures)
            _check(
                float(seoul.loc[pd.to_datetime(seoul["일시"]) == pd.Timestamp("2026-06-01"), "평균기온(°C)"].iloc[0])
                == 21.5,
                "new KMA values win on duplicate dates",
                failures,
            )
            _check(len(gimhae) == 1, "preserved sheet keeps its rows", failures)
            _check(not weather.WEATHER_EXCEL_LOCK_PATH.exists(), "lock file is removed after save", failures)

            corrupt_path = tmp_dir / "corrupt_weather.xlsx"
            corrupt_path.write_bytes(b"not-a-valid-xlsx")
            weather.PATH_WEATHER = corrupt_path
            weather.WEATHER_EXCEL_LOCK_PATH = corrupt_path.with_suffix(".xlsx.lock")
            try:
                weather._append_to_excel("서울", new_df)
                failures.append("corrupt existing weather file is not overwritten silently")
            except RuntimeError:
                pass
            _check(
                corrupt_path.read_bytes() == b"not-a-valid-xlsx",
                "corrupt existing weather file remains untouched",
                failures,
            )
        finally:
            weather.PATH_WEATHER = original_path
            weather.WEATHER_EXCEL_LOCK_PATH = original_lock

    if failures:
        print("FAIL")
        for failure in failures:
            print(f"- {failure}")
        return 1

    print("PASS")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
