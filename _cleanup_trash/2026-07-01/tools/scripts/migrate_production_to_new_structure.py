# 이 스크립트는 생산실적(완제품) 데이터를 구(舊)구조 → 신(新)구조로 전환하는
# 일회성 마이그레이션입니다. (재공품 2-파일 구조와 통일)
#
# == 구조 변화 ==
#   [구]  Raw_생산실적\ 폴더(파일별 F10_냉동.xlsx …) + DB_생산실적.xlsx(daily/월별요약)
#   [신]  RawDB_생산실적.xlsx(공장×보관유형 시트) + DB_생산실적.xlsx(공장별 wide + 제품마스터 + 계획 + daily)
#
# == 처리 단계 ==
#   1) 기존 DB_생산실적.xlsx 의 'daily' 시트(전체 이력) 로드
#   2) 기존 DB_생산실적.xlsx 백업(.bak.YYYYMMDD_HHMMSS)
#   3) daily → 신구조(공장별 wide + 제품마스터 + 계획 + daily) 로 변환 저장
#      (production_dw_service 의 빌드 헬퍼를 그대로 재사용 → 로직 일원화)
#   4) RawDB_생산실적.xlsx 생성 — 기존 Raw_생산실적\ 폴더 파일명으로 빈 카테고리 시트 구성
#      (이후 RPA 가 시트별로 채움)
#   5) 검증 — 변환 전/후 daily 의 (date,factory,item_code) 행수·actual 합계 일치 확인
#
# == DB(production_daily) 영향 ==
#   없음. 신 DB 파일도 동일 스키마의 'daily' 시트를 그대로 포함하므로
#   production_dw_sync_service 가 변경 없이 동작한다.
#
# 사용법:
#   .\.venv\Scripts\python.exe tools/scripts/migrate_production_to_new_structure.py --dry-run
#   .\.venv\Scripts\python.exe tools/scripts/migrate_production_to_new_structure.py
from __future__ import annotations

import argparse
import shutil
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook

try:
    sys.stdout.reconfigure(encoding="utf-8")  # type: ignore[attr-defined]
    sys.stderr.reconfigure(encoding="utf-8")  # type: ignore[attr-defined]
except Exception:
    pass

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))

from app.services.production_dw_service import (  # noqa: E402
    DAILY_SHEET,
    DEFAULT_OUTPUT_PATH,
    DEFAULT_RAW_PATH,
    DEFAULT_SRC_FOLDER,
    OUTPUT_COLUMNS,
    NON_FACTORY_SHEETS,
    _groups_from_names,
    _merge_factory_long,
    _merge_master,
    _merge_plan,
    load_product_groups,
    regenerate_daily,
    save_db_file,
)


def _load_legacy_daily(db_path: Path) -> pd.DataFrame:
    """기존 DB_생산실적.xlsx 의 'daily' 시트를 tidy long 으로 로드."""
    sheets = pd.ExcelFile(db_path, engine="openpyxl").sheet_names
    if DAILY_SHEET not in sheets:
        raise SystemExit(
            f"기존 파일에 '{DAILY_SHEET}' 시트가 없습니다: {db_path}\n"
            f"이미 신구조로 전환됐거나 잘못된 파일일 수 있습니다. 시트들: {sheets}"
        )
    df = pd.read_excel(db_path, sheet_name=DAILY_SHEET, engine="openpyxl")
    missing = set(OUTPUT_COLUMNS) - set(df.columns)
    if missing:
        raise SystemExit(f"'daily' 시트 필수 컬럼 누락: {missing}")
    df = df.copy()
    df["date"] = pd.to_datetime(df["date"]).dt.date
    df["item_code"] = df["item_code"].astype(str).str.replace(r"\.0$", "", regex=True)
    df["factory"] = df["factory"].astype(str)
    df["actual_qty"] = pd.to_numeric(df["actual_qty"], errors="coerce").fillna(0.0)
    df["planned_qty"] = pd.to_numeric(df["planned_qty"], errors="coerce").fillna(0.0)
    return df[OUTPUT_COLUMNS]


def _summary(df: pd.DataFrame) -> tuple[int, float]:
    """(행수, actual 합계)."""
    return len(df), round(float(df["actual_qty"].sum()), 3)


def _create_raw_file(raw_path: Path, src_folder: Path) -> list[str]:
    r"""기존 Raw_생산실적\ 폴더 파일명(stem)으로 빈 카테고리 시트 RawDB 생성.

    이후 production_daily_rpa 가 시트별로 그리드를 채운다.
    """
    if raw_path.exists():
        print(f"  RawDB 이미 존재 — 생성 생략: {raw_path}")
        return []

    stems: list[str] = []
    if src_folder.exists():
        for f in sorted(src_folder.glob("*.xlsx")):
            if f.name.startswith("~$") or f.stem.startswith("_"):
                continue
            stems.append(f.stem)

    if not stems:
        print(f"  ⚠ Raw 폴더({src_folder})에서 파일명을 못 찾음 — 빈 RawDB 생성(시트 없음).")
        print("     RPA 실행 전 RawDB_생산실적.xlsx 에 F10_냉동 등 카테고리 시트를 수동 추가하세요.")

    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거
    for sn in stems:
        if sn in NON_FACTORY_SHEETS:
            continue
        wb.create_sheet(sn)
    if not wb.sheetnames:
        wb.create_sheet("Sheet1")  # 빈 워크북 방지
    raw_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(raw_path)
    print(f"  RawDB 생성: {raw_path}  (시트 {len(stems)}개: {stems})")
    return stems


def main() -> int:
    p = argparse.ArgumentParser(description="생산실적 구→신 구조 마이그레이션")
    p.add_argument("--db", type=str, default=str(DEFAULT_OUTPUT_PATH),
                   help=f"기존/신규 DB 파일 (기본: {DEFAULT_OUTPUT_PATH})")
    p.add_argument("--raw", type=str, default=str(DEFAULT_RAW_PATH),
                   help=f"생성할 RawDB 파일 (기본: {DEFAULT_RAW_PATH})")
    p.add_argument("--src-folder", type=str, default=str(DEFAULT_SRC_FOLDER),
                   help=f"기존 Raw 폴더(시트명 도출용, 기본: {DEFAULT_SRC_FOLDER})")
    p.add_argument("--dry-run", action="store_true", help="변환 미리보기만(저장 안함)")
    args = p.parse_args()

    db_path = Path(args.db)
    raw_path = Path(args.raw)
    src_folder = Path(args.src_folder)

    print("=" * 64)
    print(" 생산실적 구→신 구조 마이그레이션")
    print(f" 시작: {datetime.now():%Y-%m-%d %H:%M:%S}  dry-run={args.dry_run}")
    print("=" * 64)

    if not db_path.exists():
        print(f"\n❌ 기존 DB 파일 없음: {db_path}")
        return 1

    # ── (1) 기존 daily 로드 ──
    print(f"\n[1/5] 기존 daily 로드: {db_path}")
    legacy = _load_legacy_daily(db_path)
    before_rows, before_sum = _summary(legacy)
    fac_before = legacy.groupby("factory").size().to_dict()
    print(f"  행수={before_rows:,}  actual합={before_sum:,.0f}")
    print(f"  factory 분포: {fac_before}")
    print(f"  날짜 범위: {legacy['date'].min()} ~ {legacy['date'].max()}")

    # ── (2) 신구조 변환 (서비스 헬퍼 재사용) ──
    print("\n[2/5] 신구조 변환 (품목군별 wide + 제품마스터 + 계획 + daily)")
    merged_long = _merge_factory_long({}, legacy)
    master_df = _merge_master({}, legacy)
    plan_df = _merge_plan(pd.DataFrame(columns=["연월", "공장", "품목코드", "계획량"]), legacy)
    new_daily = regenerate_daily(merged_long, master_df, plan_df)

    after_rows, after_sum = _summary(new_daily)
    fac_after = new_daily.groupby("factory").size().to_dict()
    print(f"  재생성 daily 행수={after_rows:,}  actual합={after_sum:,.0f}")
    print(f"  factory 분포: {fac_after}")
    print(f"  제품마스터 {len(master_df):,}품목 / 계획 {len(plan_df):,}행 / "
          f"공장시트 {list(merged_long)}")

    # ── (검증) 변환 전/후 일치 ──
    print("\n[검증] 변환 전/후 (date,factory,item_code) 집계 비교")
    ok = True
    if abs(before_sum - after_sum) > 1.0:
        print(f"  ❌ actual 합계 불일치: {before_sum:,.0f} → {after_sum:,.0f}")
        ok = False
    else:
        print(f"  ✓ actual 합계 일치 ({after_sum:,.0f})")
    # (date,factory,item_code) 단위 actual 합 비교
    g_old = legacy.groupby(["date", "factory", "item_code"])["actual_qty"].sum()
    g_new = new_daily.groupby(["date", "factory", "item_code"])["actual_qty"].sum()
    diff = g_old.subtract(g_new, fill_value=0).abs()
    n_diff = int((diff > 0.01).sum())
    if n_diff:
        print(f"  ❌ (date,factory,item) 단위 불일치 {n_diff:,}건 (상위):")
        print(diff[diff > 0.01].sort_values(ascending=False).head(10).to_string())
        ok = False
    else:
        print("  ✓ (date,factory,item_code) 단위 actual 전부 일치")

    if args.dry_run:
        print("\n[dry-run] 여기서 종료. 실제 적용 시 --dry-run 빼고 다시 실행하세요.")
        return 0 if ok else 2
    if not ok:
        print("\n❌ 검증 실패 — 저장 중단. 위 불일치를 점검하세요.")
        return 2

    # ── (3) 기존 DB 파일 백업 ──
    print("\n[3/5] 기존 DB 파일 백업")
    bak = db_path.with_name(
        db_path.stem + f".bak.{datetime.now():%Y%m%d_%H%M%S}" + db_path.suffix
    )
    shutil.copy2(db_path, bak)
    print(f"  백업: {bak.name}")

    # ── (4) 신구조 저장 (품목군 시트 = RawDB 와 동일 구성) ──
    print("\n[4/5] 신구조 DB 파일 저장 (품목군별 wide 시트)")
    # 품목군 정의: RawDB 시트명 우선, 없으면 src 폴더 파일명으로 유도
    groups = load_product_groups(raw_path)
    if not groups and src_folder.exists():
        stems = [
            f.stem for f in sorted(src_folder.glob("*.xlsx"))
            if not (f.name.startswith("~$") or f.stem.startswith("_"))
        ]
        groups = _groups_from_names(stems)
    saved = save_db_file(master_df, plan_df, new_daily, db_path, groups=groups)
    print(f"  저장: {saved}  (품목군 시트: {[g.sheet_name for g in groups]})")

    # ── (5) RawDB 생성 ──
    print("\n[5/5] RawDB_생산실적.xlsx 생성 (카테고리 시트)")
    _create_raw_file(raw_path, src_folder)

    print("\n" + "=" * 64)
    print(" ✅ 마이그레이션 완료")
    print("  - 다음 앱 기동 시 production_dw_sync_service 가 daily 시트를 그대로 동기화")
    print("  - 이후 RPA: python tools/mis_rpa/production_daily_rpa.py")
    print("=" * 64)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
